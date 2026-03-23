"""
logger.py — 模擬指標記錄與輸出。
記錄每步的 utility、social welfare、Gini、productivity、action log。
並將 agent/planner 的 thought 與記憶輸出至 Excel（.xlsx）供研究者閱讀。
"""
from __future__ import annotations

import csv
import json
import logging
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

logger = logging.getLogger(__name__)


class SimulationLogger:
    """
    記錄模擬過程中的各項指標，並在模擬結束後輸出至 CSV 和 JSON。
    """

    def __init__(self, output_dir: str = "simulation_results", run_name: str | None = None) -> None:
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.run_name = run_name or f"run_{ts}"

        self._step_logs: list[dict] = []
        self._action_logs: list[dict] = []
        self._tax_logs: list[dict] = []
        self._planner_logs: list[dict] = []
        # thought / 記憶 log（給 Excel）
        self._thought_logs: list[dict] = []  # agent 每步 thought
        self._memory_snapshots: list[dict] = []  # agent 記憶快照（長期記憶更新時）
        self._prompt_logs: list[dict] = []  # 完整 LLM prompt 記錄
        self._adjacency_events: list[dict] = []  # agent 在資源旁事件

    # ── 每步記錄 ──────────────────────────────────────────────

    def log_step(
        self,
        step: int,
        rewards: dict,
        env,
        agent_actions: dict[str, int],
        planner_action: list[int] | None = None,
    ) -> None:
        """記錄單步的指標與動作。"""
        agents = env.world.agents

        # 資產
        coin_endowments = np.array([a.total_endowment("Coin") for a in agents])
        total_coin = float(np.sum(coin_endowments))
        mean_coin = float(np.mean(coin_endowments))
        gini = _gini(coin_endowments)

        # 各 agent 效用（reward 累積）
        agent_rewards = {
            str(a.idx): float(rewards.get(str(a.idx), 0.0))
            for a in agents
        }

        # Social Welfare（Planner reward）
        planner_reward = float(rewards.get(env.world.planner.idx, 0.0))

        step_record: dict[str, Any] = {
            "step": step,
            "total_coin": total_coin,
            "mean_coin": round(mean_coin, 3),
            "gini": round(gini, 5),
            "planner_reward": round(planner_reward, 5),
            **{f"reward_agent_{i}": round(v, 5) for i, v in agent_rewards.items()},
            **{
                f"coin_agent_{str(a.idx)}": round(float(a.inventory.get("Coin", 0)), 3)
                for a in agents
            },
            **{
                f"wood_agent_{str(a.idx)}": int(a.inventory.get("Wood", 0))
                for a in agents
            },
            **{
                f"stone_agent_{str(a.idx)}": int(a.inventory.get("Stone", 0))
                for a in agents
            },
            **{
                f"labor_agent_{str(a.idx)}": round(float(a.endogenous.get("Labor", 0)), 3)
                for a in agents
            },
        }
        self._step_logs.append(step_record)

        # Action log
        action_record = {
            "step": step,
            **{f"action_agent_{k}": v for k, v in agent_actions.items()},
            "planner_action": json.dumps(planner_action) if planner_action else "NOOP",
        }
        self._action_logs.append(action_record)

        # 每 100 步打印摘要
        if step % 100 == 0:
            self._print_summary(step, mean_coin, gini, planner_reward, agent_rewards)

    def log_tax(self, step: int, tax_brackets: list[int]) -> None:
        """記錄稅率設定。"""
        self._tax_logs.append({
            "step": step,
            "tax_brackets": json.dumps(tax_brackets),
        })

    def log_planner_thought(self, step: int, thought: str, comment: str) -> None:
        """記錄 Planner 的思考過程。"""
        self._planner_logs.append({
            "step": step,
            "thought": thought,
            "society_comment": comment,
        })

    def log_thought(
        self,
        step: int,
        agent_id: str,
        agent_name: str,
        role: str,
        thought: str,
        action_id: int,
        short_term_memory: str = "",
        long_term_memory: str = "",
        is_tax_day: bool = False,
        tax_brackets: list[int] | None = None,
        society_comment: str = "",
    ) -> None:
        """
        記錄單步的 agent thought 與記憶狀態（主要供 Excel 輸出）。

        agent_id: '0'~'3' 代表 MobileAgent；'planner' 代表 Planner
        """
        self._thought_logs.append({
            "step": step,
            "agent_id": agent_id,
            "agent_name": agent_name,
            "role": role,
            "is_tax_day": is_tax_day,
            "thought": thought,
            "action_id": action_id if action_id is not None else "",
            "tax_brackets": json.dumps(tax_brackets) if tax_brackets else "",
            "society_comment": society_comment,
            "short_term_memory": short_term_memory,
            "long_term_memory": long_term_memory,
        })

    def log_memory_snapshot(
        self,
        step: int,
        agent_id: str,
        agent_name: str,
        long_term_memory: str,
        trigger: str = "consolidation",
    ) -> None:
        """記錄長期記憶更新快照。"""
        self._memory_snapshots.append({
            "step": step,
            "agent_id": agent_id,
            "agent_name": agent_name,
            "trigger": trigger,
            "long_term_memory": long_term_memory,
        })

    def log_prompt(
        self,
        step: int,
        agent_id: str,
        agent_name: str,
        system_prompt: str,
        user_prompt: str,
        context: str = "",
    ) -> None:
        """記錄送給 LLM 的完整 prompt（供診斷用）。"""
        self._prompt_logs.append({
            "step": step,
            "agent_id": agent_id,
            "agent_name": agent_name,
            "system_prompt": system_prompt,
            "user_prompt": user_prompt,
            "context": context,
        })

    def log_adjacency_event(
        self,
        step: int,
        agent_id: str,
        agent_name: str,
        resource_type: str,
        direction: str,
        agent_action: int | None = None,
        agent_thought: str = "",
    ) -> None:
        """記錄 agent 在資源旁（距離=1）的事件。"""
        self._adjacency_events.append({
            "step": step,
            "agent_id": agent_id,
            "agent_name": agent_name,
            "resource_type": resource_type,
            "direction": direction,
            "agent_action": agent_action if agent_action is not None else "",
            "agent_thought": agent_thought,
        })

    def save_map_snapshot(self, step: int, env) -> None:
        """
        輸出當步地圖快照到 {run_dir}/maps/{step}.png。
        包含：全域地圖（資源＋agent 位置）及各 agent 的 egocentric 視圖。
        """
        try:
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt
            import matplotlib.patches as mpatches
        except ImportError:
            logger.warning("[Logger] matplotlib 未安裝，跳過地圖快照")
            return

        run_dir = self.output_dir / self.run_name
        maps_dir = run_dir / "maps"
        maps_dir.mkdir(parents=True, exist_ok=True)

        # ── 地圖資料 ─────────────────────────────────────────────
        maps_state = env.world.maps.state           # (n_ch, H, W)
        channel_names: list[str] = []
        try:
            channel_names = list(env.world.maps._maps.keys())
        except AttributeError:
            channel_names = [f"ch{i}" for i in range(maps_state.shape[0])]

        W_idx  = next((i for i, n in enumerate(channel_names) if n == "Wood"), None)
        S_idx  = next((i for i, n in enumerate(channel_names) if n == "Stone"), None)
        H_idx  = next((i for i, n in enumerate(channel_names) if n in ("House", "house")), None)
        Wa_idx = next((i for i, n in enumerate(channel_names) if n == "Water"), None)

        agents = list(env.world.agents)
        n_agents = len(agents)
        AGENT_COLORS = ["#FF6B6B", "#4ECDC4", "#FFE66D", "#A29BFE",
                        "#FF8C42", "#2ECC71", "#E74C3C", "#9B59B6"]
        BG, HEADER_BG = "#0f0f23", "#1a1a2e"

        def _make_composite(state: np.ndarray) -> np.ndarray:
            comp = np.zeros((state.shape[1], state.shape[2], 3))
            if W_idx is not None:
                comp[:, :, 1] += np.clip(state[W_idx], 0, 1) * 0.75
            if S_idx is not None:
                layer = np.clip(state[S_idx], 0, 1)
                comp[:, :, 0] += layer * 0.55
                comp[:, :, 2] += layer * 0.55
            if H_idx is not None:
                comp[:, :, 0] += np.clip(state[H_idx], 0, 1)
            if Wa_idx is not None:
                comp[:, :, 2] += np.clip(state[Wa_idx], 0, 1) * 0.9
            return np.clip(comp, 0, 1)

        composite = _make_composite(maps_state)

        # ── 圖表：1 全域 + n_agents 個自我中心 ───────────────────
        ncols = 1 + n_agents
        fig, axes = plt.subplots(1, ncols, figsize=(5 * ncols, 6), facecolor=HEADER_BG)
        if ncols == 1:
            axes = [axes]

        # 全域地圖
        ax0 = axes[0]
        ax0.set_facecolor(BG)
        ax0.imshow(composite, interpolation="nearest", aspect="equal")
        for agent in agents:
            r, c = agent.loc
            color = AGENT_COLORS[agent.idx % len(AGENT_COLORS)]
            ax0.scatter(c, r, s=220, color=color, edgecolors="white", linewidths=1.2, zorder=5)
            ax0.text(c + 0.4, r - 0.6, str(agent.idx),
                     color="white", fontsize=7, fontweight="bold", zorder=6)
        ax0.set_title(f"Step {step} — Global Map", color="white", fontsize=11, pad=6)
        ax0.tick_params(colors="white", labelsize=6)
        for sp in ax0.spines.values():
            sp.set_edgecolor("#555555")
        patches = [
            mpatches.Patch(color="#00CC44", label="Wood"),
            mpatches.Patch(color="#8844BB", label="Stone"),
            mpatches.Patch(color="#FF4444", label="House"),
            mpatches.Patch(color="#2255FF", label="Water"),
        ]
        for a in agents:
            patches.append(mpatches.Patch(
                color=AGENT_COLORS[a.idx % len(AGENT_COLORS)],
                label=f"A{a.idx} ({a.loc[0]},{a.loc[1]})",
            ))
        ax0.legend(handles=patches, loc="upper right", framealpha=0.55,
                   facecolor=HEADER_BG, labelcolor="white", fontsize=6,
                   handlelength=1.0, borderpad=0.4)

        # 各 agent egocentric 視圖
        try:
            obs = env._generate_observations()
        except Exception:
            obs = {}

        for i, agent in enumerate(agents):
            ax = axes[1 + i]
            ax.set_facecolor(BG)
            agent_obs = obs.get(str(agent.idx), {})
            _vm = agent_obs.get("world-map")
            vismap = _vm if _vm is not None else agent_obs.get("map")
            if vismap is not None:
                ego = _make_composite(np.array(vismap))
                ax.imshow(ego, interpolation="nearest", aspect="equal")
                ch_c, cw_c = ego.shape[0] // 2, ego.shape[1] // 2
                ax.scatter(cw_c, ch_c, s=200,
                           color=AGENT_COLORS[agent.idx % len(AGENT_COLORS)],
                           edgecolors="white", linewidths=1.5, zorder=5, marker="*")
            else:
                ax.text(0.5, 0.5, "No ego map", transform=ax.transAxes,
                        ha="center", va="center", color="#FF6B6B", fontsize=9)

            inv = agent.inventory
            ax.set_title(
                f"A{agent.idx}  Coin:{inv.get('Coin',0):.0f} "
                f"W:{inv.get('Wood',0)} S:{inv.get('Stone',0)}",
                color=AGENT_COLORS[agent.idx % len(AGENT_COLORS)], fontsize=8, pad=4,
            )
            ax.tick_params(colors="white", labelsize=5)
            for sp in ax.spines.values():
                sp.set_edgecolor("#555555")

        plt.suptitle(f"AI Economist — Map Snapshot  Step {step}",
                     color="white", fontsize=13, y=1.01, fontweight="bold")
        plt.tight_layout(pad=1.0)
        out_path = maps_dir / f"{step}.png"
        plt.savefig(out_path, dpi=130, bbox_inches="tight",
                    facecolor=HEADER_BG, edgecolor="none")
        plt.close(fig)
        logger.info(f"[Logger] 地圖快照已儲存: {out_path}")

    # ── 輸出 ──────────────────────────────────────────────────


    def save(self) -> None:
        """將所有記錄輸出至 CSV、JSON 和 Excel。"""
        run_dir = self.output_dir / self.run_name
        run_dir.mkdir(parents=True, exist_ok=True)

        _write_csv(run_dir / "step_metrics.csv", self._step_logs)
        _write_csv(run_dir / "action_log.csv", self._action_logs)
        _write_csv(run_dir / "tax_log.csv", self._tax_logs)
        _write_csv(run_dir / "planner_thoughts.csv", self._planner_logs)

        # 整合輸出
        summary = {
            "run_name": self.run_name,
            "total_steps": len(self._step_logs),
            "final_metrics": self._step_logs[-1] if self._step_logs else {},
            "tax_count": len(self._tax_logs),
        }
        with open(run_dir / "summary.json", "w", encoding="utf-8") as f:
            json.dump(summary, f, ensure_ascii=False, indent=2)

        # Excel 輸出（thought + 記憶 + prompt + adjacency）
        if _HAS_OPENPYXL and self._thought_logs:
            xlsx_path = run_dir / "agent_thoughts.xlsx"
            _write_thoughts_excel(
                xlsx_path, self._thought_logs, self._memory_snapshots,
                self._step_logs, self._prompt_logs, self._adjacency_events,
            )
            print(f"  - agent_thoughts.xlsx ({len(self._thought_logs)} 筆 thought)")
            if self._prompt_logs:
                print(f"    └ LLM Prompts 工作表：{len(self._prompt_logs)} 筆")
            if self._adjacency_events:
                print(f"    └ 資源鄰近事件工作表：{len(self._adjacency_events)} 筆")
        elif not _HAS_OPENPYXL:
            print("  [警告] 未安裝 openpyxl，跳過 Excel 輸出。請執行：pip install openpyxl")

        print(f"\n[Logger] 結果已儲存至：{run_dir}")
        print(f"  - step_metrics.csv ({len(self._step_logs)} 筆)")
        print(f"  - action_log.csv ({len(self._action_logs)} 筆)")
        print(f"  - tax_log.csv ({len(self._tax_logs)} 筆)")
        print(f"  - summary.json")

    def _print_summary(
        self,
        step: int,
        mean_coin: float,
        gini: float,
        planner_reward: float,
        agent_rewards: dict,
    ) -> None:
        rewards_str = ", ".join(
            f"A{k}={v:.3f}" for k, v in sorted(agent_rewards.items())
        )
        print(
            f"[Step {step:>4}] "
            f"均 Coin={mean_coin:.1f} | "
            f"Gini={gini:.4f} | "
            f"Planner Reward={planner_reward:.4f} | "
            f"Agent Rewards: {rewards_str}"
        )


# ──────────────────────────────────────────────────────────────
#  工具函數
# ──────────────────────────────────────────────────────────────

def _gini(values: np.ndarray) -> float:
    values = np.sort(np.abs(values))
    n = len(values)
    if n == 0 or np.sum(values) == 0:
        return 0.0
    idx = np.arange(1, n + 1)
    return float(
        (2 * np.dot(idx, values) - (n + 1) * np.sum(values))
        / (n * np.sum(values))
    )


def _write_csv(path: Path, records: list[dict]) -> None:
    if not records:
        return
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=records[0].keys())
        writer.writeheader()
        writer.writerows(records)


def _write_thoughts_excel(
    path: Path,
    thought_logs: list[dict],
    memory_snapshots: list[dict],
    step_logs: list[dict],
    prompt_logs: list[dict] | None = None,
    adjacency_events: list[dict] | None = None,
) -> None:
    """
    輸出研究者友善的 Excel：
    Sheet 1 - Agent Thoughts：MobileAgent 每步 thought / action / 記憶
    Sheet 2 - Planner Thoughts：Planner 每步 thought / 稅率決策
    Sheet 3 - Memory Snapshots：長期記憶彙整歷史
    Sheet 4 - Step Metrics Summary：關鍵指標摘要（Coin/Gini/Reward）
    """
    if not _HAS_OPENPYXL:
        return

    wb = openpyxl.Workbook()

    # ── 通用樣式 ─────────────────────────────────────────────
    HEADER_FILL = PatternFill("solid", fgColor="2E4057")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    AGENT_FILLS = {
        "planner": PatternFill("solid", fgColor="E8F4F8"),
        "0": PatternFill("solid", fgColor="FFF8E7"),
        "1": PatternFill("solid", fgColor="F0FFF0"),
        "2": PatternFill("solid", fgColor="FFF0F0"),
        "3": PatternFill("solid", fgColor="F5F0FF"),
    }
    TAX_DAY_FONT = Font(bold=True, color="C0392B")
    WRAP = Alignment(wrap_text=True, vertical="top")
    CENTER = Alignment(horizontal="center", vertical="top")
    thin = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _safe_cell_value(value):
        """防止 Excel 將 '=' 開頭的字串誤判為公式。"""
        if isinstance(value, str) and value.startswith("="):
            return " " + value
        return value

    def _header_row(ws, headers: list[str]) -> None:
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
            cell.border = BORDER
        ws.row_dimensions[1].height = 20

    def _apply_border_wrap(ws, row_idx: int, cols: int) -> None:
        for c in range(1, cols + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.border = BORDER
            if cell.alignment.wrap_text is None:
                cell.alignment = WRAP

    # ── Sheet 1: Agent Thoughts ──────────────────────────────
    agent_logs = [r for r in thought_logs if r["agent_id"] != "planner"]
    if agent_logs:
        ws1 = wb.active
        ws1.title = "Agent Thoughts"
        headers1 = [
            "Step", "Agent ID", "Agent Name", "Role",
            "Thought（決策思維）", "Action ID",
            "短期記憶（最近N步）", "長期記憶（彙整）",
        ]
        _header_row(ws1, headers1)

        col_widths1 = [7, 9, 14, 16, 60, 10, 50, 50]
        for i, w in enumerate(col_widths1, 1):
            ws1.column_dimensions[get_column_letter(i)].width = w

        for rec in agent_logs:
            row = [
                rec["step"],
                rec["agent_id"],
                rec["agent_name"],
                _safe_cell_value(rec["role"]),
                _safe_cell_value(rec["thought"]),
                rec["action_id"],
                _safe_cell_value(rec["short_term_memory"]),
                _safe_cell_value(rec["long_term_memory"]),
            ]
            ws1.append(row)
            ri = ws1.max_row
            fill = AGENT_FILLS.get(str(rec["agent_id"]), PatternFill())
            for c in range(1, len(headers1) + 1):
                cell = ws1.cell(row=ri, column=c)
                cell.fill = fill
                cell.border = BORDER
                cell.alignment = WRAP
            ws1.row_dimensions[ri].height = 60
    else:
        ws1 = wb.active
        ws1.title = "Agent Thoughts"

    # ── Sheet 2: Planner Thoughts ────────────────────────────
    planner_logs = [r for r in thought_logs if r["agent_id"] == "planner"]
    ws2 = wb.create_sheet("Planner Thoughts")
    headers2 = [
        "Step", "Is Tax Day", "Thought（決策思維）",
        "Society Comment（社會觀察）", "Tax Brackets（稅率）",
        "短期記憶",
    ]
    _header_row(ws2, headers2)
    col_widths2 = [7, 11, 60, 50, 25, 60]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    for rec in planner_logs:
        is_tax = rec.get("is_tax_day", False)
        row = [
            rec["step"],
            "✅ 稅收日" if is_tax else "觀察",
            _safe_cell_value(rec["thought"]),
            _safe_cell_value(rec.get("society_comment", "")),
            rec.get("tax_brackets", ""),
            _safe_cell_value(rec.get("short_term_memory", "")),
        ]
        ws2.append(row)
        ri = ws2.max_row
        fill = PatternFill("solid", fgColor="FDEBD0") if is_tax else AGENT_FILLS["planner"]
        for c in range(1, len(headers2) + 1):
            cell = ws2.cell(row=ri, column=c)
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = WRAP
            if is_tax:
                cell.font = TAX_DAY_FONT
        ws2.row_dimensions[ri].height = 60

    # ── Sheet 3: Memory Snapshots ────────────────────────────
    ws3 = wb.create_sheet("Memory Snapshots")
    headers3 = ["Step", "Agent ID", "Agent Name", "Trigger", "長期記憶內容"]
    _header_row(ws3, headers3)
    col_widths3 = [7, 9, 14, 14, 80]
    for i, w in enumerate(col_widths3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    for rec in memory_snapshots:
        row = [
            rec["step"], rec["agent_id"], rec["agent_name"],
            rec["trigger"], _safe_cell_value(rec["long_term_memory"]),
        ]
        ws3.append(row)
        ri = ws3.max_row
        fill = AGENT_FILLS.get(str(rec["agent_id"]), PatternFill())
        for c in range(1, len(headers3) + 1):
            cell = ws3.cell(row=ri, column=c)
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = WRAP
        ws3.row_dimensions[ri].height = 45

    # ── Sheet 4: Step Metrics Summary ────────────────────────
    ws4 = wb.create_sheet("Step Metrics")
    if step_logs:
        # 只取關鍵欄位
        key_cols = ["step", "mean_coin", "gini", "planner_reward", "total_coin"]
        agent_reward_cols = [k for k in step_logs[0] if k.startswith("reward_agent_")]
        agent_coin_cols = [k for k in step_logs[0] if k.startswith("coin_agent_")]
        display_cols = key_cols + agent_reward_cols + agent_coin_cols
        headers4 = [c.replace("_", " ").title() for c in display_cols]
        _header_row(ws4, headers4)
        for i in range(1, len(headers4) + 1):
            ws4.column_dimensions[get_column_letter(i)].width = 14
        for rec in step_logs:
            ws4.append([rec.get(c, "") for c in display_cols])
            ri = ws4.max_row
            for c in range(1, len(headers4) + 1):
                ws4.cell(row=ri, column=c).border = BORDER
                ws4.cell(row=ri, column=c).alignment = CENTER

    # ── Sheet 5: LLM Prompts（完整提示詞記錄）────────────────
    if prompt_logs:
        ws5 = wb.create_sheet("LLM Prompts")
        headers5 = [
            "Step", "Agent ID", "Agent Name",
            "System Prompt（系統提示詞）",
            "User Prompt（觀察/狀態描述）",
            "Context（記憶背景）",
        ]
        _header_row(ws5, headers5)
        col_widths5 = [7, 9, 14, 60, 80, 50]
        for i, w in enumerate(col_widths5, 1):
            ws5.column_dimensions[get_column_letter(i)].width = w

        for rec in prompt_logs:
            row = [
                rec["step"],
                rec["agent_id"],
                rec["agent_name"],
                _safe_cell_value(rec["system_prompt"]),
                _safe_cell_value(rec["user_prompt"]),
                _safe_cell_value(rec["context"]),
            ]
            ws5.append(row)
            ri = ws5.max_row
            fill = AGENT_FILLS.get(str(rec["agent_id"]), PatternFill())
            for c in range(1, len(headers5) + 1):
                cell = ws5.cell(row=ri, column=c)
                cell.fill = fill
                cell.border = BORDER
                cell.alignment = WRAP
            ws5.row_dimensions[ri].height = 80

    # ── Sheet 6: Adjacency Events（資源鄰近事件）──────────────
    if adjacency_events:
        ws6 = wb.create_sheet("Resource Adjacency")
        headers6 = [
            "Step", "Agent ID", "Agent Name",
            "資源類型", "方向", "實際動作 ID",
            "Agent Thought（決策思維）",
        ]
        _header_row(ws6, headers6)
        col_widths6 = [7, 9, 14, 10, 20, 10, 60]
        for i, w in enumerate(col_widths6, 1):
            ws6.column_dimensions[get_column_letter(i)].width = w

        for rec in adjacency_events:
            row = [
                rec["step"],
                rec["agent_id"],
                rec["agent_name"],
                rec["resource_type"],
                rec["direction"],
                rec["agent_action"],
                _safe_cell_value(rec["agent_thought"]),
            ]
            ws6.append(row)
            ri = ws6.max_row
            fill = AGENT_FILLS.get(str(rec["agent_id"]), PatternFill())
            for c in range(1, len(headers6) + 1):
                cell = ws6.cell(row=ri, column=c)
                cell.fill = fill
                cell.border = BORDER
                cell.alignment = WRAP
            ws6.row_dimensions[ri].height = 50

    wb.save(path)
    logger.info(f"[Logger] Excel 已儲存：{path}")


def setup_logging(level: int = logging.INFO) -> None:

    """Configure global logging with UTF-8 encoding."""
    import io as _io
    root = logging.getLogger()
    root.setLevel(level)

    if not root.handlers:
        # Force UTF-8 on stdout-based handler to avoid Windows garbling
        stream = _io.TextIOWrapper(
            _io.open(sys.__stderr__.fileno(), mode="wb", closefd=False),
            encoding="utf-8",
            errors="replace",
        ) if hasattr(sys, "__stderr__") and sys.__stderr__ else sys.stderr

        handler = logging.StreamHandler(stream)
        handler.setFormatter(
            logging.Formatter(
                "%(asctime)s [%(levelname)s] %(name)s: %(message)s",
                datefmt="%H:%M:%S",
            )
        )
        root.addHandler(handler)
